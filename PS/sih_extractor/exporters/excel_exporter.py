import os
from typing import List
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..models import ProblemStatement

def export_to_excel(statements: List[ProblemStatement], output_path: str) -> str:
    """
    Exports the problem statements to an Excel workbook with formatting.
    """
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb = Workbook()
    
    # Styles
    navy_header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    title_fill = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="CCCCCC")
    
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=10)
    
    software_fill = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
    hardware_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D0D5DD'),
        right=Side(style='thin', color='D0D5DD'),
        top=Side(style='thin', color='D0D5DD'),
        bottom=Side(style='thin', color='D0D5DD')
    )
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_left_nowrap = Alignment(horizontal="left", vertical="center", wrap_text=False)
    
    columns = [
        ("S.No", 8, align_center),
        ("PS Code", 14, align_center),
        ("PS ID", 10, align_center),
        ("Category", 14, align_center),
        ("Theme", 26, align_left_nowrap),
        ("Title", 38, align_left),
        ("Organization", 32, align_left),
        ("Department", 30, align_left),
        ("Description", 60, align_left),
        ("Submissions", 14, align_center),
        ("Deadline", 18, align_center),
        ("Dataset Link", 28, align_left),
        ("YouTube Link", 28, align_left),
        ("Contact Info", 24, align_left),
    ]

    def populate_statements_sheet(ws, ps_list, sheet_title):
        ws.title = sheet_title
        ws.views.sheetView[0].showGridLines = True
        
        # Headers
        for col_idx, (col_name, col_width, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = navy_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width
            
        ws.row_dimensions[1].height = 28
        
        # Data Rows
        for r_idx, ps in enumerate(ps_list, start=2):
            row_fill = software_fill if ps.category == "Software" and sheet_title != "Software" else (
                hardware_fill if ps.category == "Hardware" and sheet_title != "Hardware" else (
                    alt_row_fill if r_idx % 2 == 0 else white_fill
                )
            )
            
            row_data = [
                ps.s_no,
                ps.code,
                ps.id,
                ps.category,
                ps.theme,
                ps.title,
                ps.organization,
                ps.department,
                ps.description,
                ps.submissions,
                ps.deadline,
                ps.dataset_link,
                ps.youtube_link,
                ps.contact_info,
            ]
            
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = columns[c_idx - 1][2]
                
                # Category highlight in 'Category' column
                if c_idx == 4:
                    if val == "Software":
                        cell.fill = software_fill
                        cell.font = Font(name="Calibri", size=10, bold=True, color="104F55")
                    elif val == "Hardware":
                        cell.fill = hardware_fill
                        cell.font = Font(name="Calibri", size=10, bold=True, color="9A3412")
                else:
                    if r_idx % 2 == 0:
                        cell.fill = alt_row_fill
                        
            ws.row_dimensions[r_idx].height = 50 # readable height for wrapped descriptions
            
        # Enable auto filters & freeze header
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(ps_list) + 1}"
        ws.freeze_panes = "A2"

    # Sheet 1: All Statements
    ws_all = wb.active
    populate_statements_sheet(ws_all, statements, "All Problem Statements")
    
    # Sheet 2: Software
    software_ps = [p for p in statements if p.category.lower() == "software"]
    if software_ps:
        ws_soft = wb.create_sheet(title="Software")
        populate_statements_sheet(ws_soft, software_ps, "Software")
        
    # Sheet 3: Hardware
    hardware_ps = [p for p in statements if p.category.lower() == "hardware"]
    if hardware_ps:
        ws_hard = wb.create_sheet(title="Hardware")
        populate_statements_sheet(ws_hard, hardware_ps, "Hardware")
        
    # Sheet 4: Summary & Analytics
    ws_sum = wb.create_sheet(title="Summary & Analytics")
    ws_sum.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_sum.merge_cells("A1:E1")
    ws_sum["A1"] = "Smart India Hackathon (SIH) 2026 - Problem Statements Summary"
    ws_sum["A1"].font = title_font
    ws_sum["A1"].fill = title_fill
    ws_sum["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_sum.row_dimensions[1].height = 36
    
    # Overview Cards
    ws_sum["A3"] = "Total Problem Statements"
    ws_sum["A3"].font = bold_font
    ws_sum["B3"] = len(statements)
    ws_sum["B3"].font = Font(name="Calibri", size=14, bold=True, color="1B365D")
    
    ws_sum["A4"] = "Software Problems"
    ws_sum["A4"].font = bold_font
    ws_sum["B4"] = len(software_ps)
    ws_sum["B4"].font = Font(name="Calibri", size=12, bold=True, color="104F55")
    
    ws_sum["A5"] = "Hardware Problems"
    ws_sum["A5"].font = bold_font
    ws_sum["B5"] = len(hardware_ps)
    ws_sum["B5"].font = Font(name="Calibri", size=12, bold=True, color="9A3412")
    
    # Category Breakdown Table
    ws_sum["A7"] = "Category Distribution"
    ws_sum["A7"].font = Font(name="Calibri", size=12, bold=True)
    ws_sum["A8"] = "Category"
    ws_sum["A8"].font = header_font
    ws_sum["A8"].fill = navy_header_fill
    ws_sum["B8"] = "Count"
    ws_sum["B8"].font = header_font
    ws_sum["B8"].fill = navy_header_fill
    ws_sum["C8"] = "Percentage"
    ws_sum["C8"].font = header_font
    ws_sum["C8"].fill = navy_header_fill
    
    cats = Counter(p.category for p in statements)
    row_cur = 9
    for cat, count in cats.most_common():
        ws_sum[f"A{row_cur}"] = cat
        ws_sum[f"B{row_cur}"] = count
        pct = count / len(statements) if statements else 0
        ws_sum[f"C{row_cur}"] = f"{pct:.1%}"
        for col_letter in ("A", "B", "C"):
            ws_sum[f"{col_letter}{row_cur}"].border = thin_border
        row_cur += 1
        
    # Theme Breakdown Table
    row_cur += 2
    ws_sum[f"A{row_cur}"] = "Theme Distribution"
    ws_sum[f"A{row_cur}"].font = Font(name="Calibri", size=12, bold=True)
    row_cur += 1
    ws_sum[f"A{row_cur}"] = "Theme"
    ws_sum[f"A{row_cur}"].font = header_font
    ws_sum[f"A{row_cur}"].fill = navy_header_fill
    ws_sum[f"B{row_cur}"] = "Count"
    ws_sum[f"B{row_cur}"].font = header_font
    ws_sum[f"B{row_cur}"].fill = navy_header_fill
    ws_sum[f"C{row_cur}"] = "Percentage"
    ws_sum[f"C{row_cur}"].font = header_font
    ws_sum[f"C{row_cur}"].fill = navy_header_fill
    
    themes = Counter(p.theme for p in statements)
    row_cur += 1
    for theme, count in themes.most_common():
        ws_sum[f"A{row_cur}"] = theme
        ws_sum[f"B{row_cur}"] = count
        pct = count / len(statements) if statements else 0
        ws_sum[f"C{row_cur}"] = f"{pct:.1%}"
        for col_letter in ("A", "B", "C"):
            ws_sum[f"{col_letter}{row_cur}"].border = thin_border
        row_cur += 1

    # Top Organizations Table
    row_cur += 2
    ws_sum[f"A{row_cur}"] = "Top Organizations / Ministries"
    ws_sum[f"A{row_cur}"].font = Font(name="Calibri", size=12, bold=True)
    row_cur += 1
    ws_sum[f"A{row_cur}"] = "Organization / Ministry"
    ws_sum[f"A{row_cur}"].font = header_font
    ws_sum[f"A{row_cur}"].fill = navy_header_fill
    ws_sum[f"B{row_cur}"] = "Problem Count"
    ws_sum[f"B{row_cur}"].font = header_font
    ws_sum[f"B{row_cur}"].fill = navy_header_fill
    
    orgs = Counter(p.organization for p in statements)
    row_cur += 1
    for org, count in orgs.most_common():
        ws_sum[f"A{row_cur}"] = org
        ws_sum[f"B{row_cur}"] = count
        for col_letter in ("A", "B"):
            ws_sum[f"{col_letter}{row_cur}"].border = thin_border
        row_cur += 1
        
    ws_sum.column_dimensions["A"].width = 45
    ws_sum.column_dimensions["B"].width = 18
    ws_sum.column_dimensions["C"].width = 18
    ws_sum.column_dimensions["D"].width = 18
    ws_sum.column_dimensions["E"].width = 18

    wb.save(output_path)
    return output_path
