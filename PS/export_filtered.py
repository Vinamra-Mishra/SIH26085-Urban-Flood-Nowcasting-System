import os
import json
import re
import shutil
import subprocess
import logging
from typing import List, Dict, Any
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sih_extractor.fetcher import fetch_html, SIH_URL
from sih_extractor.parser import parse_problem_statements, clean_text
from sih_extractor.models import ProblemStatement

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("filtered_extractor")

def evaluate_difficulty(ps: ProblemStatement) -> Dict[str, Any]:
    """
    Evaluates the difficulty score (out of 10) and technical complexity factors.
    """
    title = ps.title.lower()
    desc = ps.description.lower()
    full_text = f"{title} {desc}"
    
    # 1. Algorithmic Complexity (0 - 3.0)
    algo_score = 1.3
    algo_reasons = []
    if any(k in full_text for k in ["hydrodynamic", "physics-based", "numerical weather prediction", "nwp", "differential equation", "shallow water", "wrf"]):
        algo_score += 1.5
        algo_reasons.append("Physics-informed / Numerical simulation & fluid dynamics coupling")
    elif any(k in full_text for k in ["deep learning", "spatio-temporal", "graph neural", "transformer", "generative ai", "computer vision", "cnn", "lstm", "nowcasting"]):
        algo_score += 1.2
        algo_reasons.append("Advanced Spatio-temporal AI / Multi-modal Deep Learning models")
    elif any(k in full_text for k in ["ai/ml", "machine learning", "predictive", "anomaly detection"]):
        algo_score += 0.8
        algo_reasons.append("Predictive ML & anomaly detection models")
    else:
        algo_reasons.append("Data processing & rule-based decision engine")
        
    # 2. Data Ingestion & Engineering Complexity (0 - 3.0)
    data_score = 1.1
    data_reasons = []
    if any(k in full_text for k in ["doppler radar", "satellite imagery", "remote sensing", "multispectral", "grib", "netcdf", "dem", "elevation model", "radar"]):
        data_score += 1.5
        data_reasons.append("High-volume Doppler radar & satellite raster/GRIB2 ingestion")
    elif any(k in full_text for k in ["multi-source", "sensor data", "aws", "iot", "gis mapping", "lidar", "drone", "cadastral"]):
        data_score += 1.1
        data_reasons.append("Multi-source geospatial GIS & real-time sensor fusion")
    else:
        data_score += 0.6
        data_reasons.append("Standard API feeds & tabular time-series data")
        
    # 3. Real-Time Latency & Architecture Constraints (0 - 2.0)
    arch_score = 0.8
    arch_reasons = []
    if any(k in full_text for k in ["nowcasting", "real-time alert", "real time early warning", "sub-hourly", "0-6 hr", "immediate warning"]):
        arch_score += 1.1
        arch_reasons.append("Ultra-low-latency nowcasting & instant push alert pipelines")
    elif any(k in full_text for k in ["offline sync", "low-network", "edge", "mobile application", "field deployment"]):
        arch_score += 0.8
        arch_reasons.append("Offline-first sync & low-bandwidth field mobile apps")
    else:
        arch_score += 0.5
        arch_reasons.append("Cloud web dashboard & RESTful API microservices")
        
    # 4. Domain Rigor & Modeling Fidelity (0 - 2.0)
    domain_score = 0.8
    domain_reasons = []
    if any(k in full_text for k in ["dam break", "convective scale", "monsoon regime", "flash flood in hilly", "thermal stress", "pollution-weather coupled", "drainage and rainfall"]):
        domain_score += 1.1
        domain_reasons.append("Specialized meteorological/hydrological physics domain modeling")
    elif any(k in full_text for k in ["landslide", "urban flood", "carrying capacity", "hazard-based red zone", "relocation"]):
        domain_score += 0.8
        domain_reasons.append("Geotechnical vulnerability & multi-hazard spatial zoning")
    else:
        domain_score += 0.5
        domain_reasons.append("Disaster response planning & resource management")
        
    total_raw = algo_score + data_score + arch_score + domain_score
    final_score = round(min(9.8, max(4.0, total_raw)), 1)
    
    if final_score >= 8.5:
        level = "Extreme (Expert Level)"
        badge_color = "991B1B"
    elif final_score >= 7.5:
        level = "Very Hard (Advanced)"
        badge_color = "C2410C"
    elif final_score >= 6.5:
        level = "Hard (Intermediate-High)"
        badge_color = "D97706"
    else:
        level = "Moderate (Intermediate)"
        badge_color = "047857"
        
    # Recommended Tech Stack
    tech_stack = []
    if "radar" in full_text or "satellite" in full_text or "raster" in full_text:
        tech_stack.extend(["PyTorch / TensorFlow", "GDAL / Rasterio", "Xarray", "NetCDF4 / cfgrib"])
    elif "ai" in full_text or "ml" in full_text:
        tech_stack.extend(["PyTorch", "Scikit-Learn", "XGBoost", "Pandas"])
    if "gis" in full_text or "map" in full_text or "flood" in full_text:
        tech_stack.extend(["GeoPandas", "PostGIS", "MapLibre / Leaflet", "Deck.gl"])
    if "real-time" in full_text or "nowcast" in full_text:
        tech_stack.extend(["FastAPI", "Celery / Redis", "WebSockets"])
    else:
        tech_stack.extend(["FastAPI / Django", "PostgreSQL", "React / Next.js"])
        
    tech_stack = list(dict.fromkeys(tech_stack))
    
    return {
        "score_out_of_10": final_score,
        "difficulty_level": level,
        "badge_color": badge_color,
        "breakdown": {
            "algorithmic_complexity": round(algo_score, 1),
            "data_complexity": round(data_score, 1),
            "system_architecture": round(arch_score, 1),
            "domain_rigor": round(domain_score, 1)
        },
        "key_challenges": "; ".join(algo_reasons + data_reasons + arch_reasons + domain_reasons),
        "recommended_tech_stack": ", ".join(tech_stack[:6])
    }

def latex_escape(text: str) -> str:
    if not text:
        return ""
    s = text.replace('\\', r'\textbackslash{}')
    s = s.replace('&', r'\&')
    s = s.replace('%', r'\%')
    s = s.replace('$', r'\$')
    s = s.replace('#', r'\#')
    s = s.replace('_', r'\_')
    s = s.replace('{', r'\{')
    s = s.replace('}', r'\}')
    s = s.replace('~', r'\textasciitilde{}')
    s = s.replace('^', r'\textasciicircum{}')
    s = s.replace('<', r'\textless{}')
    s = s.replace('>', r'\textgreater{}')
    s = s.replace('°', r'$^\circ$')
    s = s.replace('±', r'$\pm$')
    s = s.replace('²', r'$^2$')
    s = s.replace('³', r'$^3$')
    s = s.replace('µ', r'$\mu$')
    s = s.replace('₹', r'Rs.~')
    s = s.replace('€', r'\texteuro{}')
    s = s.replace('©', r'\copyright{}')
    s = s.replace('®', r'\textregistered{}')
    s = s.replace('™', r'\texttrademark{}')
    s = s.replace('≥', r'$\ge$')
    s = s.replace('≤', r'$\le$')
    s = s.replace('≠', r'$\ne$')
    s = s.replace('≈', r'$\approx$')
    s = s.replace('×', r'$\times$')
    s = s.replace('÷', r'$\div$')
    s = s.replace('→', r'$\rightarrow$')
    s = s.replace('←', r'$\leftarrow$')
    return s

def format_desc_to_latex(description: str) -> str:
    if not description:
        return ""
    latex_parts = []
    lines = description.split('\n')
    
    heading_pattern = re.compile(
        r'^(Background|Description|Problem Description|Expected Solution|Key Requirements|Objectives|Scope of Work|Deliverables|The solution should|The platform should|The system should|The proposed solution should include|The proposed system should|The solution must|The platform must):?\s*(.*)',
        re.IGNORECASE
    )
    
    for line in lines:
        line_str = line.strip()
        if not line_str:
            continue
            
        hm = heading_pattern.match(line_str)
        if hm:
            heading_name = hm.group(1).strip()
            rest_text = hm.group(2).strip()
            esc_heading = latex_escape(heading_name)
            if rest_text:
                esc_rest = latex_escape(rest_text)
                latex_parts.append(r"\vspace{0.18cm}\noindent\textbf{\color{sihnavy}" + esc_heading + r":}~\small " + esc_rest + r"\par\vspace{0.06cm}")
            else:
                latex_parts.append(r"\vspace{0.18cm}\noindent\textbf{\color{sihnavy}" + esc_heading + r":}\par\vspace{0.06cm}")
            continue
            
        if line_str.startswith(('•', '-', '*')):
            content = line_str.lstrip('•-* ').strip()
            esc_content = latex_escape(content)
            latex_parts.append(r"\noindent\hspace*{1.2em}\hangindent=2.4em\hangafter=1\textcolor{sihblue}{$\bullet$}~\small " + esc_content + r"\par\vspace{0.05cm}")
            continue
            
        letter_match = re.match(r'^([a-z]\.)\s*(.*)', line_str, re.IGNORECASE)
        if letter_match:
            marker = letter_match.group(1)
            content = letter_match.group(2)
            esc_content = latex_escape(content)
            latex_parts.append(r"\noindent\hangindent=1.8em\hangafter=1\textbf{" + marker + r"}~\small " + esc_content + r"\par\vspace{0.06cm}")
            continue
            
        num_match = re.match(r'^(\d+\.)\s*(.*)', line_str)
        if num_match:
            marker = num_match.group(1)
            content = num_match.group(2)
            esc_content = latex_escape(content)
            latex_parts.append(r"\noindent\hangindent=1.8em\hangafter=1\textbf{" + marker + r"}~\small " + esc_content + r"\par\vspace{0.06cm}")
            continue
            
        esc_line = latex_escape(line_str)
        latex_parts.append(r"\noindent\small " + esc_line + r"\par\vspace{0.06cm}")
        
    return "\n".join(latex_parts)

def export_filtered():
    output_dir = os.path.abspath("filtered")
    os.makedirs(output_dir, exist_ok=True)
    
    # Clean subdirectories if any was accidentally created
    nested_dup = os.path.join(output_dir, "filtered")
    if os.path.exists(nested_dup):
        shutil.rmtree(nested_dup, ignore_errors=True)
        
    html = fetch_html(cache_file="page.html")
    all_statements = parse_problem_statements(html)
    
    # Filter for Category: Software AND Theme: Disaster Management
    filtered = []
    for ps in all_statements:
        is_software = "software" in ps.category.lower()
        is_disaster = "disaster" in ps.theme.lower()
        if is_software and is_disaster:
            diff = evaluate_difficulty(ps)
            filtered.append({
                "statement": ps,
                "difficulty": diff
            })
            
    # Sort by difficulty descending
    filtered.sort(key=lambda x: x["difficulty"]["score_out_of_10"], reverse=True)
    logger.info(f"Total Filtered Statements (Software + Disaster Management): {len(filtered)}")

    # 1. JSON Export
    json_path = os.path.join(output_dir, "sih_2026_software_disaster_management.json")
    json_payload = {
        "metadata": {
            "title": "SIH 2026 - Category: Software | Theme: Disaster Management",
            "filter_criteria": {
                "category": "Software",
                "theme": "Disaster Management"
            },
            "total_problem_statements": len(filtered),
            "difficulty_rating_scale": "1.0 - 10.0 (Evaluated across Algorithmic, Data, Architecture, and Domain complexity)",
            "average_difficulty_score": round(sum(f["difficulty"]["score_out_of_10"] for f in filtered) / len(filtered), 2) if filtered else 0
        },
        "problem_statements": [
            {
                **f["statement"].to_dict(),
                "difficulty_score": f["difficulty"]["score_out_of_10"],
                "difficulty_level": f["difficulty"]["difficulty_level"],
                "complexity_breakdown": f["difficulty"]["breakdown"],
                "key_technical_challenges": f["difficulty"]["key_challenges"],
                "recommended_tech_stack": f["difficulty"]["recommended_tech_stack"]
            }
            for f in filtered
        ]
    }
    with open(json_path, "w", encoding="utf-8") as jf:
        json.dump(json_payload, jf, indent=2, ensure_ascii=False)
    logger.info(f"Saved JSON to '{json_path}'")

    # 2. Excel Export
    excel_path = os.path.join(output_dir, "sih_2026_software_disaster_management.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Software - Disaster Mgmt"
    ws.views.sheetView[0].showGridLines = True
    
    navy_header_fill = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    
    diff_extreme_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    diff_veryhard_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
    diff_hard_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    diff_mod_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D0D5DD'),
        right=Side(style='thin', color='D0D5DD'),
        top=Side(style='thin', color='D0D5DD'),
        bottom=Side(style='thin', color='D0D5DD')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    headers = [
        ("Rank", 8, align_center),
        ("Difficulty", 14, align_center),
        ("Level", 22, align_center),
        ("PS Code", 14, align_center),
        ("PS ID", 10, align_center),
        ("Title", 36, align_left),
        ("Organization", 30, align_left),
        ("Department", 30, align_left),
        ("Key Technical Challenges", 45, align_left),
        ("Recommended Tech Stack", 32, align_left),
        ("Description", 60, align_left),
        ("Submissions", 14, align_center),
        ("Deadline", 18, align_center),
    ]
    
    for c_idx, (h_name, h_width, _) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h_name)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(c_idx)].width = h_width
        
    ws.row_dimensions[1].height = 28
    
    for r_idx, item in enumerate(filtered, start=2):
        ps = item["statement"]
        diff = item["difficulty"]
        score = diff["score_out_of_10"]
        
        diff_fill = diff_extreme_fill if score >= 8.5 else (
            diff_veryhard_fill if score >= 7.5 else (
                diff_hard_fill if score >= 6.5 else diff_mod_fill
            )
        )
        
        row_vals = [
            r_idx - 1,
            f"{score} / 10",
            diff["difficulty_level"],
            ps.code,
            ps.id,
            ps.title,
            ps.organization,
            ps.department,
            diff["key_challenges"],
            diff["recommended_tech_stack"],
            ps.description,
            ps.submissions,
            ps.deadline
        ]
        
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = bold_font if c_idx in (1, 2, 4) else regular_font
            cell.border = thin_border
            cell.alignment = headers[c_idx - 1][2]
            
            if c_idx == 2:
                cell.fill = diff_fill
                cell.font = Font(name="Calibri", size=11, bold=True, color=diff["badge_color"])
            elif c_idx == 3:
                cell.fill = diff_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color=diff["badge_color"])
            else:
                if r_idx % 2 == 0:
                    cell.fill = alt_fill
                    
        ws.row_dimensions[r_idx].height = 65
        
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(filtered) + 1}"
    ws.freeze_panes = "A2"
    wb.save(excel_path)
    logger.info(f"Saved Excel to '{excel_path}'")

    # 3. LaTeX / PDF Export via MiKTeX
    tex_path = os.path.join(output_dir, "sih_2026_software_disaster_management.tex")
    pdf_path = os.path.join(output_dir, "sih_2026_software_disaster_management.pdf")
    
    tex = []
    tex.append(r"""\documentclass[10pt,a4paper]{article}
\usepackage[utf8]{inputenc}
\usepackage[T1]{fontenc}
\usepackage{lmodern}
\usepackage[margin=1.8cm, top=2.2cm, bottom=2.2cm]{geometry}
\usepackage{xcolor}
\usepackage{titlesec}
\usepackage{fancyhdr}
\usepackage{hyperref}
\usepackage{booktabs}
\usepackage{longtable}
\usepackage{array}
\usepackage{enumitem}
\usepackage{needspace}
\usepackage[most]{tcolorbox}
\tcbuselibrary{skins,breakable}

% Color Definitions
\definecolor{sihnavy}{HTML}{0B2545}
\definecolor{sihblue}{HTML}{134074}
\definecolor{softgray}{HTML}{F8FAFC}
\definecolor{bordergray}{HTML}{CBD5E1}
\definecolor{diffextreme}{HTML}{991B1B}
\definecolor{diffveryhard}{HTML}{C2410C}
\definecolor{diffhard}{HTML}{D97706}
\definecolor{diffmod}{HTML}{047857}

\hypersetup{
    colorlinks=true,
    linkcolor=sihblue,
    urlcolor=sihblue,
    citecolor=sihblue,
    pdftitle={SIH 2026 - Software and Disaster Management Problem Statements},
    pdfauthor={SIH 2026 Extractor}
}

\pagestyle{fancy}
\fancyhf{}
\fancyhead[L]{\small\color{sihblue}\textbf{SIH 2026: Software $\vert$ Disaster Management}}
\fancyhead[R]{\small\color{gray}Difficulty Ranked Compendium}
\fancyfoot[L]{\small\color{gray}Source: sih.gov.in/sih2026PS}
\fancyfoot[R]{\small\color{sihblue}\textbf{Page \thepage}}
\renewcommand{\headrulewidth}{0.4pt}
\renewcommand{\footrulewidth}{0.4pt}

\titleformat{\section}{\Large\bfseries\color{sihnavy}}{\thesection}{1em}{}[\titlerule]
\titleformat{\subsection}{\large\bfseries\color{sihblue}}{\thesubsection}{1em}{}

\tcbset{
    pscard/.style={
        enhanced,
        breakable,
        colback=softgray,
        colframe=bordergray,
        arc=2.5mm,
        boxrule=0.7pt,
        left=9pt,
        right=9pt,
        top=9pt,
        bottom=9pt,
        fonttitle=\bfseries,
        coltitle=white,
        colbacktitle=sihnavy,
        attach boxed title to top left={yshift=-2mm, xshift=4mm},
        boxed title style={arc=1.5mm, boxrule=0.5pt}
    }
}

\begin{document}

\begin{center}
    \vspace*{0.4cm}
    {\Huge \textbf{\color{sihnavy}Smart India Hackathon 2026}}\\[0.35cm]
    {\LARGE \textbf{\color{sihblue}Category: Software \quad $\vert$ \quad Theme: Disaster Management}}\\[0.25cm]
    {\large \color{gray} Ranked by Technical Difficulty \& Complexity Score (Out of 10.0)}\\[0.4cm]
    \rule{0.85\textwidth}{1pt}\\[0.35cm]
    {\large \textbf{Total Statements: } """ + str(len(filtered)) + r""" \quad $\vert$ \quad \textbf{Difficulty Range: } 4.0 - 8.8 / 10.0}\\[0.15cm]
    {\small \color{gray} Generated on: \today}
\end{center}

\vspace{0.4cm}
\hrule
\vspace{0.4cm}

\section*{Difficulty Ranking \& Overview}
\addcontentsline{toc}{section}{Difficulty Ranking \& Overview}

\begin{tabular}{rlllp{8.5cm}}
\toprule
\textbf{\#} & \textbf{Code} & \textbf{Difficulty} & \textbf{Level} & \textbf{Title} \\
\midrule
""")
    
    for rank, item in enumerate(filtered, start=1):
        ps = item["statement"]
        diff = item["difficulty"]
        score = diff["score_out_of_10"]
        level = diff["difficulty_level"]
        tex.append(f"{rank} & \\textbf{{{latex_escape(ps.code)}}} & \\textbf{{{score}/10}} & {latex_escape(level.split()[0])} & {latex_escape(ps.title[:65])}... \\\\\n")
        
    tex.append(r"""\bottomrule
\end{tabular}

\vspace{0.8cm}
\newpage

\tableofcontents
\newpage

\section{Detailed Problem Statements}
""")
    
    for rank, item in enumerate(filtered, start=1):
        ps = item["statement"]
        diff = item["difficulty"]
        score = diff["score_out_of_10"]
        level = diff["difficulty_level"]
        short_title = ps.title[:75] + ("..." if len(ps.title) > 75 else "")
        
        diff_color = "diffextreme" if score >= 8.5 else ("diffveryhard" if score >= 7.5 else ("diffhard" if score >= 6.5 else "diffmod"))
        
        tex.append(r"\needspace{6\baselineskip}")
        tex.append(r"\phantomsection")
        tex.append(r"\addcontentsline{toc}{subsection}{" + f"[{score}/10] [{latex_escape(ps.code)}] {latex_escape(short_title)}" + r"}")
        
        formatted_desc = format_desc_to_latex(ps.description)
        
        tex.append(r"""
\begin{tcolorbox}[pscard, title={"\textbf{""" + latex_escape(ps.code) + r"""} \quad $\vert$ \quad """ + latex_escape(short_title) + r"""}]
\textbf{\large """ + latex_escape(ps.title) + r"""}\\[0.25cm]

\begin{tabular}{@{}lp{12.5cm}@{}}
\textbf{Difficulty Rating:} & \textcolor{""" + diff_color + r"""}{\textbf{\Large """ + str(score) + r""" / 10}} \quad (\textbf{""" + latex_escape(level) + r"""}) \\
\textbf{Problem ID:} & """ + latex_escape(ps.id) + r""" \quad $\vert$ \quad \textbf{Category:} \textbf{Software} \quad $\vert$ \quad \textbf{Theme:} \textbf{Disaster Management} \\
\textbf{Organization:} & """ + latex_escape(ps.organization) + r""" \\
\textbf{Department:} & """ + latex_escape(ps.department) + r""" \\
\textbf{Recommended Stack:} & \texttt{""" + latex_escape(diff["recommended_tech_stack"]) + r"""} \\
\textbf{Key Challenges:} & \textit{\small """ + latex_escape(diff["key_challenges"]) + r"""} \\
\end{tabular}

\vspace{0.2cm}
\hrule
\vspace{0.2cm}

\textbf{\color{sihnavy}Problem Statement Scope \& Requirements:}\\[0.15cm]
""" + formatted_desc + r"""
""")
        
        extras = []
        if ps.dataset_link:
            extras.append(r"\textbf{Dataset / Resources:} " + latex_escape(ps.dataset_link))
        if ps.youtube_link:
            extras.append(r"\textbf{YouTube Video:} " + latex_escape(ps.youtube_link))
        if ps.contact_info:
            extras.append(r"\textbf{Contact Info:} " + latex_escape(ps.contact_info))
            
        if extras:
            tex.append(r"""
\vspace{0.2cm}
\hrule
\vspace{0.2cm}
{\footnotesize
""" + r" \\ ".join(extras) + r"""
}
""")
            
        tex.append(r"\end{tcolorbox}")
        tex.append(r"\vspace{0.4cm}")
        
    tex.append(r"\end{document}")
    
    with open(tex_path, "w", encoding="utf-8") as tf:
        tf.write("\n".join(tex))
    logger.info(f"Saved LaTeX source to '{tex_path}'")
    
    # Compile PDF with pdflatex (Pass 1 & Pass 2 for TOC)
    pdflatex_cmd = shutil.which("pdflatex") or r"C:\Users\vkmuk\AppData\Local\Programs\MiKTeX\miktex\bin\x64\pdflatex.exe"
    logger.info(f"Compiling PDF via MiKTeX at '{pdflatex_cmd}'...")
    
    cmd = [pdflatex_cmd, "-interaction=nonstopmode", f"-output-directory={output_dir}", tex_path]
    subprocess.run(cmd, capture_output=True, text=True)
    subprocess.run(cmd, capture_output=True, text=True)
    
    for ext in [".aux", ".out", ".toc", ".log"]:
        f_to_clean = os.path.join(output_dir, f"sih_2026_software_disaster_management{ext}")
        if os.path.exists(f_to_clean):
            try:
                os.remove(f_to_clean)
            except Exception:
                pass
                
    if os.path.exists(pdf_path):
        logger.info(f"Successfully generated PDF at '{pdf_path}' ({os.path.getsize(pdf_path)/1024:.1f} KB)")
    else:
        logger.error("PDF generation failed.")
        
    print("\n[OK] Filtered Export Completed Successfully!")
    print(f"Directory: {output_dir}")
    for f in os.listdir(output_dir):
        fp = os.path.join(output_dir, f)
        if os.path.isfile(fp):
            print(f"  * {f:50s} ({os.path.getsize(fp)/1024:.1f} KB)")

if __name__ == "__main__":
    export_filtered()
